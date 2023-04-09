from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import matplotlib.pyplot as plt
import numpy as np
import gurobipy as gp
from gurobipy import GRB
from matplotlib.patches import Patch
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
import random
import os
from tkinter.filedialog import askopenfile
from openpyxl import load_workbook

N= None
J= None
R=[]
p=np.array([])
V=[]
P=[]
color = ['#E64646','#6834d0', '#d0347a','#E69646', '#34D05C',
             '#d06334', '#34D0C3', '#3475D0','#d07534', '#d09734','#d0c034',
             '#c0ff03','#79ff03','#3d5c23','#135e52','#03b3ff','#121cde',
             '#4a026b', '#ff24c5', '#ff2453']
root= Tk()
root.title('Hybrid Flow Shop')
#root['bg']='#34D0C3'

n = ttk.Notebook(root)   
n.pack()
o1 = Frame(n)

o1.pack()
o2 = ttk.Frame(n)       
o2.pack()

n.add(o1, text='Data manually')      
n.add(o2, text='Data by Excel file')

headingLabel1=Label(o1,text='Hybrid Flow Shop',
                       font=('arial',30,'bold'),
                       foreground='black')
headingLabel1.grid(row=0)
headingLabel1=Label(o2,text='Hybrid Flow Shop',
                       font=('arial',30,'bold'),
                       foreground='black')
headingLabel1.grid(row=0)
'''CheckVarxl= IntVar()
case = Checkbutton(o2, text = "Data from Excel", font=('arial', 13, 'bold'),variable = CheckVarxl,onvalue = 1, offvalue = 0, height=2,width = 25)
case.grid(row=1,pady=5)'''
def load():
    if 1==1:
        file = askopenfile(mode ='r', filetypes =[('Excel Files', '*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv')]) # To open the file that you want.
        book = load_workbook(filename = file.name) # Load into openpyxl
        sheet1=book.active
        #book=load_workbook("test.xlsx",read_only = True)
        #sheet1=book['data']
        global J
        global N
        global p
        J = sheet1.max_row #jobs
        N = sheet1.max_column #stages
        p= np.empty((N,J),int)
        for i in range(1, N+1 ):
            for j in range(1, J+1):
                p[i-1,j-1] = sheet1.cell(row=j, column=i).value
        Label(o2, text="N="+str(N)+", J="+str(J), font=('arial', 10, 'bold')).grid(row=1,column=1)
        Label(o2, text="P="+str(p), font=('arial', 10, 'bold')).grid(row=2,column=1)




load_button=Button(o2,text='Load data',
                    command=load,
                    relief='raised',
                    borderwidth=3,
                    width=25,
                    height=3,font=('arial', 13, 'bold'))
load_button.grid(row=2,pady=5)

    
def nbstagesjobs():
    if N is not None and J is not None:
        messagebox.showinfo(title=None, message="Number of stages and jobs are already defined")
    else:
        global dimension
        dimension= Tk()
        dimension.title('Number of stages and jobs')
        dimension['bg']='#34D0C3'
        nb_etage=Entry(dimension,width=20,relief="solid")
        nb_etage.grid(row=1,column=1)
        nb_etage_label=Label(dimension,text="number of stages",font=('arial', 10, 'bold'),background='#34D0C3')
        nb_etage_label.grid(row=1,column=0)
        nb_jobs=Entry(dimension,width=20,relief="solid")
        nb_jobs.grid(row=2,column=1)
        nb_jobs_label=Label(dimension,text="number of jobs",font=('arial', 10, 'bold'),background='#34D0C3')
        nb_jobs_label.grid(row=2,column=0)
        def get_NJ():
            global N
            global J
            N=int(nb_etage.get())
            J=int(nb_jobs.get())
            print(N)
            print(J)
            dimension.destroy()
            Label(o1, text="N="+str(N)+", J="+str(J), font=('arial', 10, 'bold')).grid(row=1,column=1)
        enter_btn=Button(dimension, text="submit",relief='raised',
                    borderwidth=3,command=get_NJ)
        enter_btn.grid(row=3,column=1,columnspan=1,pady=10, padx=10,ipadx=50)
        dimension.mainloop()
        
ajout_button=Button(o1,text='Number of jobs and stages',
                    command=nbstagesjobs,
                    relief='raised',
                    borderwidth=3,
                    width=25,
                    height=3,font=('arial', 13, 'bold'))
ajout_button.grid(row=1,pady=5)




def nb_machine():
    if N is None:
        messagebox.showinfo(title=None, message="Number of stage is not defined")
    else :
        global machine
        machine=Tk()
        machine.title('List mi')
        machine['bg']='#34D0C3'
        global list
        list=[]
        for i in range(N):
            Label(machine, text="m["+str(i+1)+"]", font=('arial', 10, 'bold'),background='#34D0C3').grid(row=0,column=i)
            list.append(Entry(machine,width=4))
            list[i].grid(row=1,column=i,pady=10, padx=10)
        def get_list():
            global R
            R=[]
            for i in range(N):
                R.append(int(list[i].get()))
            print(R)
            
            machine.destroy()
            Label(o1, text="m="+str(R), font=('arial', 10, 'bold')).grid(row=2,column=1)
            Label(o2, text="m="+str(R), font=('arial', 10, 'bold')).grid(row=3,column=1)
        button= Button(machine,text="Submit",relief='raised',
                    borderwidth=3, width=15, command=get_list)
        
        
        button.grid(row=2,column=N,pady=10, padx=10)
        machine.mainloop()   
    
machine_button=Button(o1,
                      text='Number of machine per stage',
                      command=nb_machine,
                      relief='raised',
                      borderwidth=3,
                      width=25,
                      height=3,font=('arial', 13, 'bold'))
machine_button.grid(row=2,pady=5)

machine_button=Button(o2,
                      text='Number of machine per stage',
                      command=nb_machine,
                      relief='raised',
                      borderwidth=3,
                      width=25,
                      height=3,font=('arial', 13, 'bold'))
machine_button.grid(row=3,pady=5)
def matrixp():
    if N is None and J is None:
        messagebox.showinfo(title=None, message="Number of stages and jobs are not defined")
    else:
        global mat
        mat=Tk()
        mat.title('matrix P ')
        mat['bg']='#34D0C3'
        global entries
        entries = []
        for i in range (N):
            Label(mat, text="Stage "+str(i+1), font=('arial', 10, 'bold'),background='#34D0C3').grid(row=0,column=i+1)
        for j in range (J):
            Label(mat, text="Job "+str(j+1), font=('arial', 10, 'bold'),background='#34D0C3').grid(row=j+1,column=0)
        for i in range(N):
            entries.append([])
            for j in range(J):
                entries[i].append(Entry(mat,width=4))
                entries[i][j].grid(row=j+1,column=i+1,pady=10, padx=10)
        def get_mat():
            global p
            p= np.empty((N,J),int)
            for i in range(N):
                for j in range(J):
                    p[i,j]=(float(entries[i][j].get()))
            print(p)
            mat.destroy()
            Label(o1, text="P="+str(p), font=('arial', 10, 'bold')).grid(row=3,column=1)
        enter_btn=Button(mat, text="Submit",relief='raised',
                    borderwidth=3,command=get_mat)
        enter_btn.grid(row=J+1,column=N+1,columnspan=1,pady=10, padx=10,ipadx=50)
        mat.mainloop()
        '''
        global p
        p= np.empty((N,J),int)
        for i in range(N):
            for j in range(J):
                p[i,j]=random.randint(1,100)
        print(p)
        '''
   
modi_button=Button(o1,
                      text='Matrix P',
                      command=matrixp,
                      relief='raised',
                      borderwidth=3,
                      width=25,
                      height=3,font=('arial', 13, 'bold'))
modi_button.grid(row=3,pady=5)




def speed():
    if N is None and not R :
        messagebox.showinfo(title=None, message="Number of stages and machines per stage are not defined")
    elif p.size==0:
        messagebox.showinfo(title=None, message="Matrix p is empty")
    else:
        global sp
        sp=Tk()
        sp.title('Speeds of machines ')
        sp['bg']='#34D0C3'
        max_value = None
        for num in R:
            if (max_value is None or num > max_value):
                max_value = num
        global entries1
        entries1 = []
        for i in range (N):
            Label(sp, text="Stage "+str(i+1), font=('arial', 10, 'bold'),background='#34D0C3').grid(row=0,column=i+1)
        for l in range (max_value):
            Label(sp, text="Machine "+str(l+1), font=('arial', 10, 'bold'),background='#34D0C3').grid(row=l+1,column=0)
        for i in range(N):
            entries1.append([])
            for l in range(R[i]):
                entries1[i].append(Entry(sp,width=4))
                entries1[i][l].grid(row=l+1,column=i+1,pady=10, padx=10)
        def get_sp():
            for i in range(N):
                V.append([round(float(entries1[i][l].get()),1) for l in range (R[i])])

            for i in range (N):
                for l in range (R[i]):
                    print("V["+str(i)+"]["+str(l)+"]= "+str(V[i][l]))
           
            for i in range (N):
                P.append([])
                for l in range(R[i]):
                    P[i].append([int(p[i,j]/(V[i][l]*0.01)) for j in range (J)])
            for i in range (N):
                for l in range(R[i]):
                    for j in range (J):
                        print("P["+str(i)+"]["+str(l)+"]["+str(j)+"]= "+str(P[i][l][j]))
        
            sp.destroy()
            Label(o1, text="V="+str(V), font=('arial', 10, 'bold')).grid(row=4,column=1)
            Label(o2, text="V="+str(V), font=('arial', 10, 'bold')).grid(row=4,column=1)
        enter_btn=Button(sp, text="Submit",relief='raised',
                    borderwidth=3,command=get_sp)
        enter_btn.grid(row=max_value+1,column=N+1,columnspan=1,pady=10, padx=10,ipadx=50)
        sp.mainloop()

speed_button=Button(o1,
                      text='Speeds of machines',
                      command=speed,
                      relief='raised',
                      borderwidth=3,
                      width=25,
                      height=3,font=('arial', 13, 'bold'))
speed_button.grid(row=4,pady=5)



speed_button=Button(o2,
                      text='Speeds of machines',
                      command=speed,
                      relief='raised',
                      borderwidth=3,
                      width=25,
                      height=3,font=('arial', 13, 'bold'))
speed_button.grid(row=4,pady=5)


def resoudre():
    
    global m
    global Z
    global T
    global d
    global W
    global Y
    M=10000
    m=gp.Model('HFS')
    T=m.addVars(N,J, vtype=GRB.INTEGER, name="T")
    d=m.addVars(N,J, vtype=GRB.INTEGER,name="d")
    W=[]
    for i in range (N):
        W.append(m.addVars(J,R[i],lb=0,ub=1, vtype=GRB.BINARY, name="W[{}]".format(i)))
    Y=m.addVars(N,J,J, lb=0,ub=1,vtype=GRB.BINARY, name="Y")
    
    Z=m.addVar(vtype=GRB.INTEGER)
    if v.get()==1:
        m.setObjective(1*Z, GRB.MINIMIZE)
        for j in range(J):
            for l in range(R[N-1]):
                m.addConstr(1*T[N-1,j]+P[N-1][l][j]*(W[N-1])[j,l]<=Z)
    elif v.get()==2 :
        m.setObjective(sum(1*T[N-1,j]+p[i,j]-T[0,j] for j in range (J) )/J, GRB.MINIMIZE)
    elif v.get()==3:
        L=m.addVars(J, vtype=GRB.CONTINUOUS, name="L")
        Tr=m.addVars(J,lb=0, vtype=GRB.CONTINUOUS, name="T")
        for j in range (J):
            m.addConstr(L[j]==T[N-1,j]+p[N-1,j]-du[j])
        for j in range (J):
            m.addGenConstrMax(Tr[j],[L[j]],0.0)
        m.addGenConstrMax(Z,[Tr[j] for j in range (J)])
        m.setObjective(Z,GRB.MINIMIZE)
       # m.setObjective(sum(Tr[j] for j in range(J))/J, GRB.MINIMIZE)

    if CheckVar1.get()==1 and CheckVar2.get()==1:
        for j in range(J):
            m.addConstr(r[j]<=T[0,j])
            m.addConstr(T[N-1,j]+p[N-1,j]<=du[j])
    elif CheckVar1.get()==1:
        for j in range(J):
            m.addConstr(r[j]<=T[0,j])
    elif CheckVar2.get()==1:
        for j in range(J):
            m.addConstr(T[N-1,j]+p[N-1,j]<=du[j])
    if CheckVar3.get()==1:
        for i in range(N-1):
            for l in range(R[i]):
                for j in range(J):
                    for k in range(J):
                        if j!=k:
                            m.addConstr(1*T[i,j]+P[i][l][j]+1*d[i,j]-M*(1-Y[i,j,k])-M*(1-(W[i])[j,l])-M*(1-(W[i])[k,l])<=T[i,k],"C1")
    else:
        for i in range(N-1):
            for l in range(R[i]):
                for j in range(J):
                    for k in range(J):
                        if j!=k:
                            m.addConstr(1*T[i,j]+P[i][l][j]-M*(1-Y[i,j,k])-M*(1-(W[i])[j,l])-M*(1-(W[i])[k,l])<=T[i,k],"C1")
    for j in range(J):
        for l in range(R[N-1]):
            for k  in range(J):
                if j!=k:
                    m.addConstr(1*T[N-1,j]+P[N-1][l][j]-M*(1-Y[N-1,j,k])-M*(1-(W[N-1])[j,l])-M*(1-(W[N-1])[k,l])<=T[N-1,k],"C12")
    for i in range(N):
        for j in range(J):
            for k in range(J):
                if j!=k:
                    m.addConstr(1*Y[i,j,k]+1*Y[i,k,j]==1,"C4")
    if CheckVar3.get()==1:
        for i in range(N-1):
            for j in range(J):
                for l in range(R[i]):
                    m.addConstr(1*T[i,j]+P[i][l][j]+1*d[i,j]==T[i+1,j],"C3")
    else :
        for i in range(N-1):
            for j in range(J):
                for l in range(R[i]):
                    m.addConstr(T[i+1,j]-T[i,j]>=P[i][l][j],"C3")
    for i in range(N) :
        for j in range(J):
            m.addConstr(sum((W[i])[j,l] for l in range(R[i]))==1,"C2")
    if len(gap.get())!=0:
        m.params.MIPGap =int(gap.get())*0.01
    
    if len(time.get())!=0:
        m.params.TimeLimit = int(time.get())
    
    m.optimize()
   
    if m.Status == GRB.INFEASIBLE:
        print("model is infeasible11")
        constraint=Label(conf,text="Model is infeasible",font=('arial', 10, 'bold'),background='#34D0C3')
        constraint.grid(row=8,column=0)
    else :
        '''
        print("les valeurs des variables de décisions")
        for i in range(N):
            for j in range(J):
                for l in range(R[i]):
                    print("W["+str(i+1)+","+str(j+1)+","+str(l+1)+"]="+str(m.getVarByName("W["+str(i)+"]["+str(j)+","+str(l)+"]").X))
        for i in range(N):
            for j in range(J):
                for k in range(J):
                    if j!=k:
                        print("Y["+str(i+1)+","+str(j+1)+","+str(k+1)+"]="+str(m.getVarByName("Y["+str(i)+","+str(j)+","+str(k)+"]").X))
        for i in range(N):
            for j in range(J):
                print("T["+str(i+1)+","+str(j+1)+"]="+str(round(m.getVarByName("T["+str(i)+","+str(j)+"]").X,1)))
        for i in range(N-1):
            for j in range(J):
                print("d["+str(i+1)+","+str(j+1)+"]="+str(round(m.getVarByName("d["+str(i)+","+str(j)+"]").X,1)))
        '''        

        print("la valeur de la fonction objectif")
        print(round(m.ObjVal,1))
        constraint=Label(conf,text="objective="+str(round(m.ObjVal,1)),font=('arial', 10, 'bold'),background='#34D0C3')
        constraint.grid(row=8,column=0)
        wb = load_workbook("resultat.xlsx")
        now = datetime.now()
        sheetname =now.strftime("%Y.%m.%d-%H.%M.%S")
        wb.create_sheet(index = 0 ,title = sheetname)
        sheet = wb[sheetname]
        dataxl=[]
        txl=("Job","machine","T","P")
        dataxl.append(txl)
        for j in range (J):
            for i in range(N):
                for l in range (R[i]):
                    if (m.getVarByName("W["+str(i)+"]["+str(j)+","+str(l)+"]").X==1):
                        txl=("J"+str(j+1),"M"+str(i+1)+"."+str(l+1),m.getVarByName("T["+str(i)+","+str(j)+"]").X,P[i][l][j])
                        dataxl.append(txl)
                    else :
                        txl=("J"+str(j+1),"M"+str(i+1)+"."+str(l+1),0,0)
                        dataxl.append(txl)
        for i in dataxl:
            sheet.append(i)
        wb.save("resultat.xlsx")
        graph()
    
def rjdj():
    global rd
    rd= Tk()
    rd.title('Constraint Settings')
    rd['bg']='#34D0C3'
    global listd
    global listr
    if  CheckVar1.get()==1 and CheckVar2.get()==1:
        print("rd")
        listd=[]
        listr=[]
        for j in range(J):
            Label(rd, text="r["+str(j+1)+"]", font=('arial', 10, 'bold'),background='#34D0C3').grid(row=j,column=0)
            listr.append(Entry(rd,width=4))
            listr[j].grid(row=j,column=1,pady=10, padx=10)
            Label(rd, text="d["+str(j+1)+"]", font=('arial', 10, 'bold'),background='#34D0C3').grid(row=j,column=2)
            listd.append(Entry(rd,width=4))
            listd[j].grid(row=j,column=3,pady=10, padx=10)
    elif CheckVar1.get()==1:
        print("r")
        listr=[]
        for j in range(J):
            Label(rd, text="r["+str(j+1)+"]", font=('arial', 10, 'bold'),background='#34D0C3').grid(row=j,column=0)
            listr.append(Entry(rd,width=4))
            listr[j].grid(row=j,column=1,pady=10, padx=10)
    elif CheckVar2.get()==1:
        print("d")
        listd=[]
        for j in range(J):
            Label(rd, text="d["+str(j+1)+"]", font=('arial', 10, 'bold'),background='#34D0C3').grid(row=j,column=0)
            listd.append(Entry(rd,width=4))
            listd[j].grid(row=j,column=1,pady=10, padx=10)
    def get_rd():
        global du
        global r
        if CheckVar2.get()==1 and CheckVar1.get()==1:
            du=[]
            r=[]
            for j in range(J):
                r.append(float(listr[j].get()))
                du.append(float(listd[j].get()))
        elif CheckVar1.get()==1:
            r=[]
            for j in range(J):
                r.append(float(listr[j].get()))
        elif CheckVar2.get()==1:
            du=[]
            for j in range(J):
                du.append(float(listd[j].get()))
        
        rd.destroy()
    getrd_btn=Button(rd, text="Submit",
                      borderwidth=3,font=('arial', 10, 'bold'),command=get_rd)
    getrd_btn.grid(row=J,column=3,columnspan=1,pady=10, padx=10,ipadx=50)
    rd.mainloop()

def configuration():
    if not R:
        messagebox.showinfo(title=None, message="Number of machines per stage is not defined")
    elif N is None and J is None:
        messagebox.showinfo(title=None, message="Number of stages and jobs are not defined")
    elif p.size==0:
        messagebox.showinfo(title=None, message="Matrix p is empty")
    elif not V:
        for i in range (N):
                P.append([])
                for l in range(R[i]):
                    P[i].append([p[i,j] for j in range (J)])
        '''for i in range (N):
                for l in range(R[i]):
                    for j in range (J):
                        print("P["+str(i)+"]["+str(l)+"]["+str(j)+"]= "+str(P[i][l][j]))
'''
    if not R:
        messagebox.showinfo(title=None, message="Number of machines per stage is not defined")
    elif N is None and J is None:
        messagebox.showinfo(title=None, message="Number of stages and jobs are not defined")
    elif p.size==0:
        messagebox.showinfo(title=None, message="Matrix p is empty")
    else :
        global conf
        conf= Tk()
        conf.title('Resolve')
        conf['bg']='#34D0C3'
        obj=Label(conf,text="objective",font=('arial', 10, 'bold'),background='#34D0C3')
        obj.grid(row=3,column=0)
        def sel():
            selection=Label(conf,text="You selected the option " + str( v.get()),font=('arial', 10, 'bold'),background='#34D0C3')
            selection.grid(row=7,column=0)
        global v
        v = IntVar(conf,10)
        case1=Radiobutton(conf, text = "Makespan",variable = v, value=1 ,background='#34D0C3',command=sel)
        case1.grid(row=4,column=0)
        case2=Radiobutton(conf, text = "Mean flow time",variable = v, value=2,background='#34D0C3',command=sel)
        case2.grid(row=4,column=1)
        case3=Radiobutton(conf, text = "Maximum Lateness",variable = v, value=3,background='#34D0C3',command=sel)
        case3.grid(row=4,column=2)
        constraint=Label(conf,text="Constraint",font=('arial', 10, 'bold'),background='#34D0C3')
        constraint.grid(row=0,column=0)
        global CheckVar3
        CheckVar3=IntVar(conf)
        case3 = Checkbutton(conf, text = "With blocking", variable = CheckVar3,onvalue = 1, offvalue = 0,background='#34D0C3')
        case3.grid(row=1,column=2,sticky=W)
        global CheckVar1
        global CheckVar2
        CheckVar1= IntVar(conf)
        case4 = Checkbutton(conf, text = "Release dates", variable = CheckVar1,onvalue = 1, offvalue = 0,background='#34D0C3')
        case4.grid(row=1,column=0,sticky=W)
        CheckVar2= IntVar(conf)
        case5 = Checkbutton(conf, text = "Due dates", variable = CheckVar2,onvalue = 1, offvalue = 0,background='#34D0C3')
        case5.grid(row=1,column=1,sticky=W)
        global gap
        gap=Entry(conf,width=8,relief="solid")
        gap.grid(row=5,column=1)
        gap_label=Label(conf,text="Gap (%)",font=('arial', 10, 'bold'),background='#34D0C3')
        gap_label.grid(row=5,column=0)
        global time
        time=Entry(conf,width=8,relief="solid")
        time.grid(row=6,column=1)
        time_label=Label(conf,text="Time (s)",font=('arial', 10, 'bold'),background='#34D0C3')
        time_label.grid(row=6,column=0)
        conf_btn=Button(conf, text="Constraint setting",relief='raised',
                    borderwidth=3,command=rjdj)
        conf_btn.grid(row=2,column=1,columnspan=1,pady=10, padx=10,ipadx=50)
        resoudre_btn=Button(conf, text="Resolve",relief='raised',
                    borderwidth=3,command=resoudre)
        resoudre_btn.grid(row=9,column=1,columnspan=1,pady=10, padx=10,ipadx=50)
        graph_btn=Button(conf, text="Gantt chart",relief='raised',
                    borderwidth=3,command=graph)
        graph_btn.grid(row=10,column=1,columnspan=1,pady=10, padx=10,ipadx=50)
        root.destroy()
        
        conf.mainloop()



resolve_button=Button(o1,
                      text='submit the input',
                      command=configuration,
                      relief='raised',bg ='blue',
                      borderwidth=3,
                      width=20,
                      height=3,font=('arial', 13, 'bold'))
resolve_button.grid(row=6,pady=5)

resolve_button=Button(o2,
                      text='submit the input',
                      command=configuration,
                      relief='raised',bg ='blue',
                      borderwidth=3,
                      width=22,
                      height=3,font=('arial', 13, 'bold'))
resolve_button.grid(row=5,pady=5)

        
           
                
 


    
def graph():
    if m.Status==GRB.OPTIMAL or m.Status==GRB.SOLUTION_LIMIT:
       
        fig, ax = plt.subplots(1, figsize=(16,6))
        plt.grid(axis="y",which="major",ls="--", lw=1)
        plt.grid(axis="x",which="major",ls="--", lw=1)
        ax.set_xlabel('time in min')
        ax.set_ylabel('Machine')
        for j in range (J):
            for i in range(N):
                for l in range (R[i]):
                    if (m.getVarByName("W["+str(i)+"]["+str(j)+","+str(l)+"]").X==1):
                        t=("M"+str(i+1)+"."+str(l+1),round(m.getVarByName("T["+str(i)+","+str(j)+"]").X,1),P[i][l][j])
                        ax.barh(str(t[0]),t[2],left=t[1],color=color[j])
                        ax.text(round(m.getVarByName("T["+str(i)+","+str(j)+"]").X,1)+P[i][l][j],"M"+str(i+1)+"."+str(l+1),str(round(m.getVarByName("T["+str(i)+","+str(j)+"]").X,1)+P[i][l][j]), va='center', alpha=0.8)
                    else :
                        t=("M"+str(i+1)+"."+str(l+1),0,0)
                        ax.barh(str(t[0]),t[2],left=t[1],color=color[j])
        legend_elements = [Patch(facecolor=color[j], label="j"+str(j+1))  for j in range(J)]
        plt.legend(handles=legend_elements)
        plt.title("Objective ="+str(round(m.ObjVal,1)))
        plt.show()
    else :
         messagebox.showinfo(title=None, message="Problem is not resolved")
        
'''graph_button=Button(o1,
                      text='Gantt chart',
                      command=graph,
                      relief='raised',
                      borderwidth=3,
                      width=25,
                      height=3,font=('arial', 13, 'bold'))
graph_button.grid(row=6,pady=5)
graph_button=Button(o2,
                      text='Gantt chart',
                      command=graph,
                      relief='raised',
                      borderwidth=3,
                      width=25,
                      height=3,font=('arial', 13, 'bold'))
graph_button.grid(row=6,pady=5)
'''
root.mainloop()
